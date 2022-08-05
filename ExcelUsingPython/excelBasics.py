from openpyxl import Workbook

#Using the openpyxl package to create an excel sheet
#Creating a workbook on excel using this line below.
wb = Workbook()
#a workbook is always created with one worksheet. To get the first worksheet, use the line below.
ws0 = wb.active
#We can create new worksheets using the line of code below. This will add the sheet created to the end of the list of sheets.
ws1 = wb.create_sheet("MySheet")
#We can also creaete a sheet and give it a particular position to be placed in. The line below adds the sheet to the first poisiton in the list of sheets.
ws2 = wb.create_sheet("firstSheet",0)
#You can change the name of sheets using the fowwolowing line.
ws0.title = "sheet1"
#You can get the names of all the sheets using the following line. Gives the names of the sheet in an array.
print(wb.sheetnames)
#You can loop through worksheets with the line below. This will print the sheet names on individual lines.
for sheet in wb:
    print(sheet.title)
#Creating copies of a worksheet in the same workbook. The name of the copy sheet is the same as the source sheet, but with the word copy added to the end.
source = wb.active
target = wb.copy_worksheet(source)
#You cannot copy worksheets between workbooks.
#We are now going to access cells in a worksheet. To access a cell in a sheet, using the following command.
c = ws0['A4']
#Adding a value to a cell
ws0['A4'] = "Divyesh"
#When a worksheet is created in memory, it is not created with cells.
#To access a cell range, you need to use the colon operator.
cell_range = ws0['A1':'C2']
#Adding data to the cells in the xcell sheet
ws0["A1"] = "Hello"
ws1["B1"] = "World!"
#Deleting a sheet from the workbook
wb.remove(target)
wb.remove(ws1)
wb.remove(ws2)
#To save a file
wb.save(filename = "Hello_world.xlsx")
